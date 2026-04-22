import io

from openpyxl import load_workbook

from app.parsers.base import CandidateItem, DocumentParser, ParsedDocument, build_fragment_ref


class XlsxParser(DocumentParser):
    def parse(self, source: str | bytes) -> ParsedDocument:
        wb = (
            load_workbook(io.BytesIO(source), data_only=True)
            if isinstance(source, bytes)
            else load_workbook(source, data_only=True)
        )
        candidates: list[CandidateItem] = []
        raw_text_chunks: list[str] = []
        fragment_idx = 1

        for sheet in wb.worksheets:
            previous_row_text = ""
            header_row: list[str] | None = None
            for row in sheet.iter_rows(values_only=True):
                values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if not values:
                    continue
                if header_row is None:
                    header_row = values
                row_text = " | ".join(values)
                raw_text_chunks.append(row_text)
                row_values: dict[str, str] = {}
                if header_row and values != header_row:
                    width = min(len(header_row), len(values))
                    for idx in range(width):
                        header = header_row[idx].strip()
                        value = values[idx].strip()
                        if header and value:
                            row_values[header] = value
                candidates.append(
                    CandidateItem(
                        item_type="sheet_row",
                        title="Spreadsheet fragment",
                        content=row_text,
                        source_ref=build_fragment_ref(fragment_idx),
                        confidence=None,
                        extracted_data={
                            "fragment_kind": "spreadsheet",
                            "section_title": sheet.title,
                            "parent_section_titles": [],
                            "nearby_text": previous_row_text,
                            "table_context": {
                                "is_structured_table": True,
                                "column_headers": header_row or [],
                                "row_values": row_values,
                            },
                        },
                    )
                )
                previous_row_text = row_text
                fragment_idx += 1

        return ParsedDocument(raw_text="\n".join(raw_text_chunks), candidates=candidates)
